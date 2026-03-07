"""Temporary file export registry for CSV/XLSX data downloads."""

from __future__ import annotations

import csv
import io
import re
import time
import uuid
from dataclasses import dataclass, field
from typing import Any


@dataclass
class ExportedFile:
    """Metadata for a generated export file."""

    file_id: str
    filename: str
    format: str
    content: bytes
    created_at: float
    title: str | None = None
    row_count: int = 0


# Default TTL: 1 hour
_TTL_SECONDS = 3600

# Guard against DoS via huge exports
MAX_EXPORT_ROWS = 10_000
MAX_EXPORT_COLUMNS = 200


@dataclass
class ExportRegistry:
    """In-memory store for exported files with TTL eviction."""

    _files: dict[str, ExportedFile] = field(default_factory=dict)

    def _evict(self) -> None:
        cutoff = time.time() - _TTL_SECONDS
        expired = [fid for fid, f in self._files.items() if f.created_at < cutoff]
        for fid in expired:
            del self._files[fid]

    def get(self, file_id: str) -> ExportedFile | None:
        self._evict()
        return self._files.get(file_id)

    def store(self, ef: ExportedFile) -> None:
        self._evict()
        self._files[ef.file_id] = ef


# Singleton registry
export_registry = ExportRegistry()


def generate_csv(
    columns: list[str],
    rows: list[list[Any]],
    title: str | None = None,
) -> bytes:
    """Generate a CSV file as bytes."""
    buf = io.StringIO()
    if title:
        buf.write(f"# {title}\n")
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compat


def generate_xlsx(
    columns: list[str],
    rows: list[list[Any]],
    title: str | None = None,
) -> bytes:
    """Generate an Excel workbook as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Export")[:31]  # Excel sheet name limit

    start_row = 1

    # Title row
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        start_row = 3  # skip a blank row

    # Header row
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_value(value))

    # Auto-width columns
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(columns[col_idx - 1]))
        for row_idx in range(start_row + 1, start_row + 1 + len(rows)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce_value(value: Any) -> Any:
    """Try to convert string values to numbers for Excel.

    Preserves strings with leading zeros (e.g. "007") and non-numeric
    patterns to avoid data corruption.
    """
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return value
    # Preserve leading zeros (sample IDs, codes)
    if len(stripped) > 1 and stripped[0] == "0" and stripped[1] != ".":
        return value
    # Only coerce strings that look like plain numbers (no scientific notation)
    if re.match(r"^-?\d+$", stripped):
        return int(stripped)
    if re.match(r"^-?\d+\.\d+$", stripped):
        return float(stripped)
    return value


def create_export(
    columns: list[str],
    rows: list[list[Any]],
    fmt: str,
    filename: str | None = None,
    title: str | None = None,
) -> ExportedFile:
    """Create an exported file and register it for download."""
    if len(rows) > MAX_EXPORT_ROWS:
        raise ValueError(f"Export limited to {MAX_EXPORT_ROWS} rows (got {len(rows)}).")
    if len(columns) > MAX_EXPORT_COLUMNS:
        raise ValueError(f"Export limited to {MAX_EXPORT_COLUMNS} columns (got {len(columns)}).")

    file_id = uuid.uuid4().hex[:12]
    base = filename or "export"
    # Sanitize filename — strip anything that isn't alphanumeric, dash, underscore, or space
    base = "".join(c for c in base if c.isalnum() or c in "-_ ").strip()
    if not base:
        base = "export"

    if fmt == "xlsx":
        content = generate_xlsx(columns, rows, title=title)
        full_filename = f"{base}.xlsx"
    elif fmt == "csv":
        content = generate_csv(columns, rows, title=title)
        full_filename = f"{base}.csv"
    else:
        # Placeholder for custom template format
        content = generate_csv(columns, rows, title=title)
        full_filename = f"{base}.csv"

    ef = ExportedFile(
        file_id=file_id,
        filename=full_filename,
        format=fmt,
        content=content,
        created_at=time.time(),
        title=title,
        row_count=len(rows),
    )
    export_registry.store(ef)
    return ef
